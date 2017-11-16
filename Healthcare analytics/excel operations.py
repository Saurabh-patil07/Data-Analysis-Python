#____________________________________________________________________________________

#- Author name: Saurabh Patil                              
#- Last updated: 10/15/2017
#- Title : Excel operations for healthcare analytics
#- Code version: v0.01
#- Type: Python source code
#
#- © All rights are reserved.
# ___________________________________________________________________________________


import pandas
import requests
import openpyxl
import sqlite3


def getHospitalInfomation(p_Id):
	conn = sqlite3.connect("medicare_hospital_compare.db")
	c = conn.cursor()
	Id = (int(p_Id),)
	c.execute("SELECT hospital_name,city,state,county_name FROM hospital_general_information WHERE provider_id=?",Id)
	hospital = c.fetchone()
	conn.close()
	return hospital

def CreateStatewiseDataFrame(state_abrevation):
	index= ['provider_id','state']
	state_df = pandas.DataFrame(columns=index)
	conn = sqlite3.connect("medicare_hospital_compare.db")
	for i in df['Provider ID']:
		c = conn.cursor();
		c.execute("SELECT provider_id,state FROM hospital_general_information WHERE provider_id=? AND state=?",(int(i), state_abrevation))
		state_hospitals = c.fetchone()
		if state_hospitals != None:
			state_df= state_df.append(pandas.Series(state_hospitals[0:2], index=index), ignore_index=True)
		if len(state_df.index) == 100:
			break
	conn.close()
	return state_df

#Download excel workbook
url = "http://kevincrook.com/utd/hospital_ranking_focus_states.xlsx"	
r = requests.get(url)
xf = open("hospital_ranking_focus_states.xlsx","wb")
xf.write(r.content)
xf.close()
df = pandas.read_excel("hospital_ranking_focus_states.xlsx", sheetname="Hospital National Ranking")
focus_list = pandas.read_excel("hospital_ranking_focus_states.xlsx", sheetname="Focus States")
focus_list = focus_list.sort_values(['State Name'])
df = df.sort_values(['Ranking'], ascending=True)
new_df = df.head(n=100)

#Create new excel workbook
wb = openpyxl.Workbook()
sheet_1 = wb.create_sheet("Nationwide")
sheet_1.cell(row = 1, column = 1, value = "Provider ID")
sheet_1.cell(row = 1, column = 2, value = "Hospital Name")
sheet_1.cell(row = 1, column = 3, value = "City")
sheet_1.cell(row = 1, column = 4, value = "State")
sheet_1.cell(row = 1, column = 5, value = "County")
row_no=2
for provider_Id in new_df['Provider ID']:
	sheet_1.cell(row = row_no, column = 1, value = provider_Id)
	hospital = getHospitalInfomation(provider_Id)
	sheet_1.cell(row = row_no, column = 2, value = hospital[0])
	sheet_1.cell(row = row_no, column = 3, value = hospital[1])
	sheet_1.cell(row = row_no, column = 4, value = hospital[2])
	sheet_1.cell(row = row_no, column = 5, value = hospital[3])
	row_no += 1

for index, row in focus_list.iterrows():
	state_name = row['State Name']
	state_name = wb.create_sheet(state_name)
	state_name.cell(row = 1, column = 1, value = "Provider ID")
	state_name.cell(row = 1, column = 2, value = "Hospital Name")
	state_name.cell(row = 1, column = 3, value = "City")
	state_name.cell(row = 1, column = 4, value = "State")
	state_name.cell(row = 1, column = 5, value = "County")
	state_df = CreateStatewiseDataFrame(row['State Abbreviation'])
	row_no=2
	print(state_df)
	for provider_Id in state_df['provider_id']:
		state_name.cell(row = row_no, column = 1, value = provider_Id)
		hospital = getHospitalInfomation(provider_Id)
		state_name.cell(row = row_no, column = 2, value = hospital[0])
		state_name.cell(row = row_no, column = 3, value = hospital[1])
		state_name.cell(row = row_no, column = 4, value = hospital[2])
		state_name.cell(row = row_no, column = 5, value = hospital[3])
		row_no += 1


wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
wb.save("hospital_ranking.xlsx")
