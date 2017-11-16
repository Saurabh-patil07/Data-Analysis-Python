#____________________________________________________________________________________

#- Author name: Saurabh Patil                              
#- Last updated: 10/27/2017
#- Title : Healthcare analytics for Hospital recommendation
#- Code version: v0.02
#- Type: Python source code
#
#- © All rights are reserved.
# ___________________________________________________________________________________


import requests
import os
import zipfile
import sqlite3
import pandas
import glob
import openpyxl
import numpy as np


def GetHospitalInfomation(p_Id):
	conn = sqlite3.connect("medicare_hospital_compare.db")
	c = conn.cursor()
	Id = (int(p_Id),)
	c.execute("SELECT hospital_name,city,state,county_name FROM hospital_general_information WHERE provider_id=?",Id)
	hospital = c.fetchone()
	conn.close();
	return hospital

def CreateStatewiseHealthDf(state_abrevation):
	index= ['provider_id','state']
	state_df = pandas.DataFrame(columns=index)
	conn = sqlite3.connect("medicare_hospital_compare.db")
	for i in df['Provider ID']:
		c = conn.cursor();
		c.execute("SELECT provider_id,state FROM hospital_general_information WHERE provider_id=? AND state=?",(int(i), state_abrevation))
		state_hospitals = c.fetchone()
		if state_hospitals != None:
			state_df= state_df.append(pandas.Series(state_hospitals[0:2], index=index), ignore_index=True)
		if len(state_df.index) == 100:
			break
	conn.close();
	return state_df

def CalculateScoreNationwide(measure_id):
	index= ['measure_id','state', 'score']
	score_df = pandas.DataFrame(columns=index)
	conn = sqlite3.connect("medicare_hospital_compare.db")
	c = conn.cursor();
	c.execute("SELECT measure_id,state,score FROM timely_and_effective_care___hospital WHERE measure_id=?",(measure_id,))
	measure_score = c.fetchall()
	for scores in measure_score:
		score_df= score_df.append(pandas.Series(scores[0:3], index=index), ignore_index=True)
	conn.close();
	new_score = score_df[score_df.score != "Not Available"]
	new_score.loc[:,'score']= new_score.loc[:,'score'].replace(['Low (0 - 19,999 patients annually)'], 9999.5)
	new_score.loc[:,'score']= new_score,loc[:,'score'].replace(['Medium (20,000 - 39,999 patients annually)'], 29999.5)
	new_score.loc[:,'score']= new_score.loc[:,'score'].replace(['High (40,000 - 59,999 patients annually)'], 49999.5)
	new_score.loc[:,'score']= new_score.loc[:,'score'].replace(['Very High (60,000+ patients annually)'], 60000)
	scores = list(map(float, new_score["score"].tolist()))
	del new_score; del score_df
	score = []
	score.append(max(scores))
	score.append(min(scores))
	score.append((sum(scores)/len(scores)))
	score.append(np.std(scores))
	return score

def CalculateScoreStatewise(measure_id, state_abrevation):
	index= ['measure_id','state', 'score']
	score_df = pandas.DataFrame(columns=index)
	conn = sqlite3.connect("medicare_hospital_compare.db")
	c = conn.cursor();
	c.execute("SELECT measure_id,state,score FROM timely_and_effective_care___hospital WHERE measure_id=? AND state=?",(measure_id, state_abrevation))
	measure_score = c.fetchall()
	for scores in measure_score:
		score_df= score_df.append(pandas.Series(scores[0:3], index=index), ignore_index=True)
	conn.close();
	new_score = score_df[score_df.score != "Not Available"]
	new_score.loc[:,'score']= new_score.loc[:,'score'].replace(['Low (0 - 19,999 patients annually)'], 9999.5)
	new_score.loc[:,'score']= new_score,loc[:,'score'].replace(['Medium (20,000 - 39,999 patients annually)'], 29999.5)
	new_score.loc[:,'score']= new_score.loc[:,'score'].replace(['High (40,000 - 59,999 patients annually)'], 49999.5)
	new_score.loc[:,'score']= new_score.loc[:,'score'].replace(['Very High (60,000+ patients annually)'], 60000)
	scores = list(map(float, new_score["score"].tolist()))
	del new_score; del score_df
	score = []
	score.append(min(scores))
	score.append(max(scores))
	score.append(np.mean(scores))
	score.append(np.std(scores))
	return score


#-----------------------------------Create Staging subdirectory and extract zip file-------------------------------------

staging_dir_name = "staging"
os.mkdir(staging_dir_name)
os.path.isdir(staging_dir_name)
url="https://data.medicare.gov/views/bg9k-emty/files/0a9879e0-3312-4719-a1db-39fd114890f1?content_type=application%2Fzip%3B%20charset%3Dbinary&filename=Hospital_Revised_Flatfiles.zip"
r = requests.get(url)
r.headers
zip_file_name = os.path.join(staging_dir_name, "Hospital_Revised_Flatfiles.zip")
zf = open(zip_file_name, "wb")
zf.write(r.content)
zf.close()
z = zipfile.ZipFile(zip_file_name,"r")
z.extractall(staging_dir_name)
z.close()

#--------------------------------------Create 'medicare_hospital_compare.db' database file-------------------------------
conn = sqlite3.connect("medicare_hospital_compare.db")

for file_name in glob.glob(staging_dir_name+'\*.csv'):
	if "FY2015_Percent_Change_in_Medicare_Payments.csv" in file_name: continue
	df = pandas.read_csv(file_name, encoding = "ISO-8859-1")
	df = df.rename(columns=lambda x: x.lower())
	df = df.rename(columns=lambda x: x.replace(' ', '_'))
	df = df.rename(columns=lambda x: x.replace('-', '_'))
	df = df.rename(columns=lambda x: x.replace('%', 'pct'))
	df = df.rename(columns=lambda x: x.replace('/', '_'))
	df = df.rename(columns=lambda x: "c_"+x if (not x[0].isalpha()) else x)
	file_name = file_name.strip('staging\\')
	table_name = file_name.strip('.csv')
	table_name = table_name.lower()
	table_name = table_name.replace(' ', '_')
	table_name = table_name.replace('-', '_')
	table_name = table_name.replace('%', 'pct')
	table_name = table_name.replace('/', '_')
	if (not table_name[0].isalpha()): table_name = "t_"+table_name
	df.to_sql(table_name, conn, if_exists='append', index=False)
conn.close()

#-----------------------------Download 'hospital_ranking_focus_states.xlsx' workbook-----------------------------------

url = "http://kevincrook.com/utd/hospital_ranking_focus_states.xlsx"	
r = requests.get(url)
xf = open("hospital_ranking_focus_states.xlsx","wb")
xf.write(r.content)
xf.close()
df = pandas.read_excel("hospital_ranking_focus_states.xlsx", sheetname="Hospital National Ranking")
focus_list = pandas.read_excel("hospital_ranking_focus_states.xlsx", sheetname="Focus States")
focus_list = focus_list.sort_values(['State Name'])
df = df.sort_values(['Ranking'], ascending=True)
new_df = df.head(n=100)

#-----------------------------Create the Hospital Ranking MS Excel Workbook------------------------------------------
wb = openpyxl.Workbook()

#Create the Nationwide Hospital Ranking MS Excel Worksheet
sheet_1 = wb.create_sheet("Nationwide")
sheet_1.cell(row = 1, column = 1, value = "Provider ID")
sheet_1.cell(row = 1, column = 2, value = "Hospital Name")
sheet_1.cell(row = 1, column = 3, value = "City")
sheet_1.cell(row = 1, column = 4, value = "State")
sheet_1.cell(row = 1, column = 5, value = "County")
#Populate data into Nationwide Hospital Ranking MS Excel Worksheet
row_no=2
for provider_Id in new_df['Provider ID']:
	sheet_1.cell(row = row_no, column = 1, value = provider_Id)
	hospital = GetHospitalInfomation(provider_Id)
	sheet_1.cell(row = row_no, column = 2, value = hospital[0])
	sheet_1.cell(row = row_no, column = 3, value = hospital[1])
	sheet_1.cell(row = row_no, column = 4, value = hospital[2])
	sheet_1.cell(row = row_no, column = 5, value = hospital[3])
	del hospital
	row_no += 1
#Create Statewise Hospital Ranking MS Excel Worksheet
for index, row in focus_list.iterrows():
	state_name = row['State Name']
	state_name = wb.create_sheet(state_name)
	state_name.cell(row = 1, column = 1, value = "Provider ID")
	state_name.cell(row = 1, column = 2, value = "Hospital Name")
	state_name.cell(row = 1, column = 3, value = "City")
	state_name.cell(row = 1, column = 4, value = "State")
	state_name.cell(row = 1, column = 5, value = "County")
	state_df = CreateStatewiseHealthDf(row['State Abbreviation'])
	#Populate data into Statewise Hospital Ranking MS Excel Worksheet
	row_no=2
	for provider_Id in state_df['provider_id']:
		state_name.cell(row = row_no, column = 1, value = provider_Id)
		hospital = GetHospitalInfomation(provider_Id)
		state_name.cell(row = row_no, column = 2, value = hospital[0])
		state_name.cell(row = row_no, column = 3, value = hospital[1])
		state_name.cell(row = row_no, column = 4, value = hospital[2])
		state_name.cell(row = row_no, column = 5, value = hospital[3])
		del hospital
		row_no += 1

wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
wb.save("hospital_ranking.xlsx")

#----------------------------------Create the Measures Statistical Analysis MS Excel Workbook-----------------------------

index= ['measure_id', 'measure_name']
measure_df = pandas.DataFrame(columns=index)
conn = sqlite3.connect("medicare_hospital_compare.db")
c = conn.cursor();
c.execute("SELECT DISTINCT measure_id,measure_name FROM timely_and_effective_care___hospital")
measures = c.fetchall()
for measure in measures:
    measure_df= measure_df.append(pandas.Series(measure[0:2], index=index), ignore_index=True)
conn.close()
measure_df = measure_df.sort_values(['measure_id'], ascending=True)
#Create Nationwide Measure Statstics excel worksheet
wb = openpyxl.Workbook()
sheet_1 = wb.create_sheet("Nationwide")
sheet_1.cell(row = 1, column = 1, value = "Measure ID")
sheet_1.cell(row = 1, column = 2, value = "Measure Name")
sheet_1.cell(row = 1, column = 3, value = "Minimum")
sheet_1.cell(row = 1, column = 4, value = "Maximum")
sheet_1.cell(row = 1, column = 5, value = "Average")
sheet_1.cell(row = 1, column = 6, value = "Standard Deviation")
#Populate data into Nationwide Measure Statstics excel worksheet
row_no=2
for index, row in measure_df.iterrows():
	sheet_1.cell(row = row_no, column = 1, value = row['measure_id'])
	sheet_1.cell(row = row_no, column = 2, value = row['measure_name'])
	#score = CalculateScoreNationwide(row['measure_id'])
	#sheet_1.cell(row = row_no, column = 3, value = score[0])
	#sheet_1.cell(row = row_no, column = 4, value = score[1])
	#sheet_1.cell(row = row_no, column = 5, value = score[2])
	#sheet_1.cell(row = row_no, column = 6, value = score[3])
	row_no += 1
del measure_df
#Create Statewise Measure Statstics excel worksheet
for index, row in focus_list.iterrows():
	state_name = row['State Name']; print(state_name)
	index= ['measure_id', 'measure_name','state']
	measure_state_df = pandas.DataFrame(columns=index)
	conn = sqlite3.connect("medicare_hospital_compare.db")
	c = conn.cursor();
	c.execute("SELECT DISTINCT measure_id,measure_name,state FROM timely_and_effective_care___hospital WHERE state=?",(row['State Abbreviation'],))
	measures = c.fetchall()
	for measure in measures:
	   	measure_state_df= measure_state_df.append(pandas.Series(measure[0:3], index=index), ignore_index=True)
	conn.close()
	measure_state_df = measure_state_df.sort_values(['measure_id'], ascending=True)
	state_name = wb.create_sheet(state_name)
	state_name.cell(row = 1, column = 1, value = "Measure ID")
	state_name.cell(row = 1, column = 2, value = "Measure Name")
	state_name.cell(row = 1, column = 3, value = "Minimum")
	state_name.cell(row = 1, column = 4, value = "Maximum")
	state_name.cell(row = 1, column = 5, value = "Average")
	state_name.cell(row = 1, column = 6, value = "Standard Deviation")
	#Populate data into Statewise Measure Statstics excel worksheet
	row_no=2
	for index, row in measure_state_df.iterrows():
		state_name.cell(row = row_no, column = 1, value = row['measure_id'])
		state_name.cell(row = row_no, column = 2, value = row['measure_name'])
		#score = CalculateScoreNationwide(row['measure_id'])
		#state_name.cell(row = row_no, column = 3, value = score[0])
		#state_name.cell(row = row_no, column = 4, value = score[1])
		#state_name.cell(row = row_no, column = 5, value = score[2])
		#state_name.cell(row = row_no, column = 6, value = score[3])
		row_no += 1
	del measure_state_df
wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
wb.save("measure_statistics.xlsx")